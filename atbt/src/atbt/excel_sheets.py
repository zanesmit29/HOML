# openpyxl can work for both Excel and LibreOffice '.xlsx' format
# if using cloud Excel download a copy run the Python file to edit it and re-upload
# if in desktop close the excel first and then run python
# in Libre office save it as 'Excel 2007–365' file type

import openpyxl
import openpyxl.chart
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl.workbook
import pprint
# wb = openpyxl.load_workbook('C:/Users/zanes/Projects/atbt/src/atbt/example3.xlsx')
# print(type(wb))

# Getting the sheets
# sheet = wb['Sheet3']
# print(sheet)
# print(sheet.title)
# another_sheet = wb.active #get the active sheet
# print(another_sheet)

# Getting cells from the sheet
# sheet = wb['Sheet1']
# cell_1 = sheet['A1']
# print(cell_1)
# print(cell_1.value)

# # Getting the row and column values
# cell_2 = sheet['B1']
# print(f'Row {cell_2.row}, Column {cell_2.column} is value {cell_2.value}')

# Go through every other row
# sheet = wb['Sheet1']
# # c = sheet.cell(row=1, column= 2).value
# for i in range(1, 8, 2):
#     print(sheet.cell(row=i, column = 2).value)

# Find the maximums
# sheet = wb['Sheet1']
# print(sheet.max_row)
# print(sheet.max_column)

# Converting between column letters and numbers
# print(get_column_letter(1))
# print(get_column_letter(27))
# sheet = wb['Sheet1']
# print(get_column_letter(sheet.max_column))

# Getting rows and columns
# sheet = wb['Sheet1']
# for row_of_cell_objects in sheet['A1':'C3']:
#     for cell_obj in row_of_cell_objects:
#         print(cell_obj.coordinate, cell_obj.value)
#     print('---END OF ROW---')

# # list(sheet.columns)[1] ## Gets the second column's cells
# for cell_obj in list(sheet.columns)[1]:
#     print(cell_obj.value)


#Project
wb = openpyxl.load_workbook('C:/Users/zanes/Projects/atbt/src/atbt/censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
county_data = {}

for row in range (2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    county_data.setdefault(state, {})
    county_data[state].setdefault(county, {'tracts': 0, 'pop': 0})
    county_data[state][county]['tracts'] += 1
    county_data[state][county]['pop'] += int(pop)

print('Writing the results....')
result_file = open('census2010.py', 'w')
result_file.write('allData = ' + pprint.pformat(county_data)) #pformat produces a string that is valid python code
result_file.close()
print('Done')

# Creating and saving excel files
wb1 = openpyxl.Workbook() # Creates a blank workbook
sheet = wb1.active
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet' # Changes the title of the sheet
wb1.create_sheet(index = 0, title = 'First sheet') # Creates a new sheet at a certain index and name
print(wb1.sheetnames)

# Writing values to a cell
sheet = wb1['First sheet']
sheet['A1'] = 'Hello, World!'
print(sheet['A1'].value)

# Formulas
wb2 = openpyxl.Workbook()
sheet5 = wb2['Sheet']
sheet5['A1'] = 200
sheet5['A2'] = 300
sheet5['A3'] = '=SUM(A1:A2)'
wb2.save('writeFormula3.xlsx')

# Setting row height and column
wb3 = openpyxl.Workbook()
sheet6 = wb3['Sheet']
sheet6['A1'] = 'Tall row'
sheet6['B2'] = 'Wide column'
sheet6.row_dimensions[1].height = 70
sheet6.column_dimensions['B'].width = 20
wb3.save('dimensions3.xlsx')

# Charts
wb4 = openpyxl.Workbook()
sheet7 = wb4['Sheet']
for i in range(1, 11):
    sheet7['A' + str(i)] = i * i

ref_obj = openpyxl.chart.Reference(sheet7, 1, 1, 1, 10)
series_obj = openpyxl.chart.Series(ref_obj, title='First series')
chart_obj = openpyxl.chart.BarChart()
chart_obj.title = 'My Chart'
chart_obj.append(series_obj)
sheet7.add_chart(chart_obj, 'C5')
wb4.save('sampleChart3.xlsx')



